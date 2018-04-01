from sys import argv
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import sys

def getData(data):

    data_matrix = [[row[0].value,row[1].value,row[2].value, x] for x,row in enumerate(data.iter_rows(min_row=1)) if type(row[0].value) == float]
    data_matrix = np.array(data_matrix)
    separation = []
    #print data_matrix
    for x,i in enumerate(data_matrix[1:,3]):
        if i != (data_matrix[x,3]+1):
            separation.append(x+1)

    data_matrix = data_matrix[:,0:3]
    data_matrix = np.split(data_matrix, separation)
    return data_matrix

# WORKING
def calc_diff_and_norm(m,c):
    #print m.shape
    normalized = []
    differences = []

    for i in range(2*c):
        diff = np.subtract(m[i][:,2],m[i+(2*c)][:,2])
        differences.append(diff)
    #print differences
    for norm in differences:
        g = np.subtract(norm, norm.min())
        normalized.append(g)
    #print normalized

    return differences, normalized


def get_maximum(n):
    max_com = [x.argmax() for x in n]
    return max_com

def get_array_boundaries(pos, n):
    min_dist_from_center = min(pos)
    #print pos
    #print n[0][0].shape[0]
    dist_center_end = [(x.shape[0])-pos[i] for i,x in enumerate(n)]
    #print dist_center_end
    max_dist_from_center = min(dist_center_end)
    #print "THIS IS MIN AND MAX"
    #print min_dist_from_center
    #print max_dist_from_center
    return (min_dist_from_center,max_dist_from_center)

def get_sets(n, aLen, c_pos, c):
    set1 = np.zeros((aLen[1]+aLen[0],0))
    set2 = set1.copy()
    #print c
    #print len(n)

    for g,i in enumerate(n[0:c]):
        to_stack1 = i[c_pos[g]-aLen[0]:c_pos[g]+aLen[1]]
        to_stack1 = np.reshape(to_stack1, (to_stack1.shape[0],1))
        #to_stack1 = to_stack1.T
        #print to_stack1.shape
        #print set1
        set1 = np.hstack([set1,to_stack1])

    for f,j in enumerate(n[c:2*c]):
        to_stack2 = j[c_pos[f]-aLen[0]:c_pos[f]+aLen[1]]
        to_stack2 = np.reshape(to_stack2, (to_stack2.shape[0],1))
        set2 = np.hstack([set2,to_stack2])

    #print "sets ------"
    #print set1
    #print set2

    # get averages
    avgs1 = np.mean(set1, axis = 1)
    #print avgs1
    avgs2 = np.mean(set2, axis = 1)
    #print avgs2
    return avgs1,avgs2

def clean_list(norm, pass_check, comets):
    #print pass_check
    set1_values = [norm_val for idx, norm_val in enumerate(norm[0:comets]) if (idx in pass_check)]
    #print "////////////////////////////////"
    #print set1_values
    #print "////////////////////////////////"
    set2_values = [norm_val for idx, norm_val in enumerate(norm[comets:]) if (idx in pass_check)]
    #print set2_values
    acceppted_normalized_matrix = np.append(set1_values,set2_values)
    #print pass_check
    #print acceppted_normalized_matrix

    
    #print set2_values
    return acceppted_normalized_matrix

def print_results(set1, set2):
    print "-------------Averages-------------"
    print "Set1 ------------------------ Set2"
    print "----------------------------------"
    for i,x in enumerate(set1):
        print "{0:.8f}            {1:.8f}".format(set1[i],set2[i])
    print "----------------------------------"

def save_results(wb,avg_set_1, avg_set_2, n, pc):
    # put results into new sheet
    wb1 = wb.create_sheet(title="Results")
    # write Titles
    wb1['A1'] = "Normalized Values"
    wb1['B1'] = "# of Chosen Comets"
    wb1['D1'] = "Averages"
    wb1['D2'] = "Set 1"
    wb1['E2'] = "Set 2"

    # write data Normalized Values
    idx = 3
    for i,set_d in enumerate(n):
        wb1['A{}'.format(idx)] = "Set {}".format(i)
        idx += 1
        for m,j in enumerate(set_d):
            wb1['A{}'.format(idx)] = j
            idx += 1
        idx = m

    # write data # of Chosen Comets
    wb1['B3'] = len(pc)

    # write data Averages of sets
    for i,x in enumerate(avg_set_1):
        wb1['D{}'.format(i+3)] = avg_set_1[i]
        wb1['E{}'.format(i+3)] = avg_set_2[i]

    return
"""Collect command-line options in a dictionary"""

def getopts(argv):
    opts = {}  # Empty dictionary to store key-value pairs.
    while argv:  # While there are arguments left to parse...
        if argv[0][0] == '-':  # Found a "-name value" pair.
            opts[argv[0]] = argv[1]  # Add key and value to the dictionary.
        argv = argv[1:]  # Reduce the argument list by copying it starting from index 1.
    return opts

if __name__ == '__main__':

    myargs = getopts(argv)

    ## INITIALIZATION
    # load filename
    file =  myargs['-i']
    # load workbook
    wb = load_workbook(filename = file)
    # retrieve number of comets
    comets = int(myargs['-c'])
    # retrieve first sheet name
    data = wb.active
    # obtain matrices by iterating over it
    matrices = getData(data)
    
    # PREPARATION
    # calculate the differences and the normalized differences
    diff, norm = calc_diff_and_norm(matrices,comets)
    # check if set fullfills constraints
    pass_check = [x for x,v in enumerate(norm[0:comets]) if v.max() > 100 and v.max() < 250]
    #print norm
    #print np.add(pass_check,1)
    # clean list
    final_norm = clean_list(norm, pass_check, comets)
    #print final_norm

    comets = len(pass_check)
    #print comets
    print "-------------"
    print final_norm
    print "-------------"
    #print pass_check
    #if len(pass_check) != len(diff):
    #    print "sorry but this dataset cannot be regarded"
    #    sys.exit()

    # get highest values from the different commets for set 1
    pos_of_max_vals = get_maximum(final_norm[0:comets])
    #print pos_of_max_vals
    # check the array lengths
    array_lengths = get_array_boundaries(pos_of_max_vals, final_norm[0:comets])
    # take averages of set 1 & set 2
    averages_set_1, averages_set_2 = get_sets(final_norm, array_lengths, pos_of_max_vals, comets)

    # give result
    print_results(averages_set_1, averages_set_2)
    save_results(wb, averages_set_1, averages_set_2, norm, pass_check)

    