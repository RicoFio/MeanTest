from sys import argv
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

def getData(data):

    data_separation = np.zeros((0,3))

    for idx, row in enumerate(data.iter_rows(min_row=1)):
        new_data = np.array([row[0].value,row[1].value,row[2].value])
        if type(new_data[0]) == np.float64:
    	    #print new_data
            data_separation = np.vstack([data_separation,new_data])
    	    #print "added"
        
    return data_separation

def calc_diff_and_norm(m,dp,c):
    #print m.shape
    normalized = np.zeros((db,0))
    difference = (np.subtract(m[:,0],m[:,1])).T

    for i in range((c*2)-1):
        norm = difference[i*dp:(i+1)*dp,:]
        norm = np.subtract(norm, norm.min())
        normalized = np.hstack((normalized, norm))

    return difference, normalized

def gen_mean_matrix(mm, idx, diff):
    # first we look how much space there is above and below the max value

    return

"""Collect command-line options in a dictionary"""

def getopts(argv):
    opts = {}  # Empty dictionary to store key-value pairs.
    while argv:  # While there are arguments left to parse...
        if argv[0][0] == '-':  # Found a "-name value" pair.
            opts[argv[0]] = argv[1]  # Add key and value to the dictionary.
        argv = argv[1:]  # Reduce the argument list by copying it starting from index 1.
    return opts

if __name__ == '__main__':

    myargs = getopts(argv)

    # load filename
    file =  myargs['-i']
    # load workbook
    wb = load_workbook(filename = file)
    # retrieve number of comets
    comets = int(myargs['-c'])
    # retrieve first sheet name
    data = wb.active
    # obtain matrices by iterating over it
    matrices = getData(data)
    # calculate differences and the normalize differences
    difference, normalized = calc_diff_and_norm(matrices, data_points, comets)
    # indices of min val
    indices = [np.argmax(v) for v in normalized.T]
    # discard data if necessary
    max_val_first_group = np.argmax(normalized[:,0:comets])
    if max_val_first_group < 100 or max_val_first_group > 250 :
        print "Sorry, this data set cannot be regarded."
        return
    # generate mean matrix
    mean_matrix = normalized.copy()
    mean_matrix = gen_mean_matrix(mean_matrix, indices, normalized)


    #print matrices[1:size]
